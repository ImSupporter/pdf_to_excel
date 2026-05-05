import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")


def _write_sheet(ws, headers: list[str], rows: list[dict]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row_idx, col, row.get(header, ""))


def _collect_headers(rows: list[dict]) -> list[str]:
    headers: list[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)
    return headers


def export_to_excel(
    broker_raw: dict[str, list[dict]],
    output_path: str,
) -> None:
    """
    broker_raw: {"증권사명": [원본_행_dict, ...]}
    output_path: 저장할 .xlsx 경로
    증권사별로 시트 1개씩 생성. 빈 증권사는 건너뜀.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 Sheet 제거

    for broker_name, raw_rows in broker_raw.items():
        if not raw_rows:
            continue
        ws = wb.create_sheet(title=broker_name[:31])  # Excel 시트명 31자 제한
        headers = _collect_headers(raw_rows)
        _write_sheet(ws, headers, raw_rows)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="결과없음")

    wb.save(output_path)
