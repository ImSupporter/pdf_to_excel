from dataclasses import dataclass
from pathlib import Path

import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.models import STANDARD_FIELDS

YELLOW_FILL = "FFFF00"
GRAY_FILL = "BFBFBF"
META_SHEET = "_metadata"
FIELDS_SHEET = "필드목록"

_AUTO_DATE_PATTERNS: list[tuple[str, str]] = [
    (r"\d{4}/\d{2}/\d{2}", "yyyy/mm/dd"),
    (r"\d{4}-\d{2}-\d{2}", "yyyy-mm-dd"),
    (r"\d{4}\.\d{2}\.\d{2}", "yyyy.mm.dd"),
    (r"\d{2}/\d{2}/\d{4}", "dd/mm/yyyy"),
]


def date_format_to_re(fmt: str) -> str:
    r"""Convert user-friendly date format string to regex. e.g. 'yyyy/mm/dd' → r'\d{4}/\d{2}/\d{2}'"""
    result = fmt
    result = result.replace("yyyy", r"\d{4}")
    result = result.replace("yy", r"\d{2}")
    result = result.replace("mm", r"\d{2}")
    result = result.replace("dd", r"\d{2}")
    result = result.replace(".", r"\.")
    return result


def _detect_date_format(texts: list[str]) -> tuple[str, str] | None:
    """Scan texts for a known date pattern. Returns (regex, format_str) or None."""
    import re
    for pattern, fmt in _AUTO_DATE_PATTERNS:
        compiled = re.compile(pattern)
        if any(compiled.match(t) for t in texts):
            return pattern, fmt
    return None


@dataclass
class TemplateCell:
    page_index: int
    row_index: int
    column_index: int
    x: float
    y: float
    text: str


@dataclass
class AnnotatedField:
    standard_field: str
    row_offset: int
    x: float


@dataclass
class TemplateAnnotations:
    field_mappings: list[AnnotatedField]
    skip_keywords: list[str]
    detected_date_format: str | None = None


def _cell_fill_rgb(cell) -> str:
    fill = cell.fill
    if fill.fill_type != "solid":
        return ""
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb[-6:].upper()
    if color.type == "indexed":
        indexed = {
            5: "FFFF00",
            22: "C0C0C0",
            23: "808080",
        }
        return indexed.get(color.indexed, "")
    return ""


def is_yellow(cell) -> bool:
    rgb = _cell_fill_rgb(cell)
    return rgb in {"FFFF00", "FFF2CC", "FFD966"}


def is_gray(cell) -> bool:
    rgb = _cell_fill_rgb(cell)
    return rgb in {"BFBFBF", "C0C0C0", "D9D9D9", "808080", "A6A6A6"}


def _normalize_label(text: str) -> str:
    return "".join(ch for ch in str(text).lower() if ch.isalnum())


def infer_standard_field(text: str) -> str | None:
    aliases = {
        "date": ["date", "거래일자", "일자", "거래일", "매매일자"],
        "type": ["type", "거래종류", "구분", "매매구분", "거래구분"],
        "ticker": ["ticker", "종목코드", "코드", "symbol"],
        "name": ["name", "종목명", "종목", "상품명"],
        "quantity": ["quantity", "수량", "거래수량", "잔고수량"],
        "price": ["price", "단가", "가격", "체결단가"],
        "amount": ["amount", "거래금액", "금액", "매매금액"],
        "fee": ["fee", "수수료", "수수료금액"],
        "tax": ["tax", "세금", "거래세", "제세금"],
        "balance": ["balance", "잔액", "잔고", "평가금액"],
        "broker": ["broker", "증권사"],
    }
    normalized = _normalize_label(text)
    for key, values in aliases.items():
        labels = values + [STANDARD_FIELDS.get(key, "")]
        if normalized in {_normalize_label(v) for v in labels if v}:
            return key
    return None


def _compute_x_zones(cells: list[TemplateCell], x_gap: float = 20.0) -> list[float]:
    """Cluster x-coordinates across all cells into distinct column zones."""
    xs = sorted({c.x for c in cells})
    if not xs:
        return [0.0]
    zones = [xs[0]]
    for x in xs[1:]:
        if x - zones[-1] > x_gap:
            zones.append(x)
    return zones


def _find_zone_index(x: float, zones: list[float]) -> int:
    """Return the index of the zone nearest to x."""
    return min(range(len(zones)), key=lambda i: abs(zones[i] - x))


def _extract_page_cells(page: fitz.Page, page_index: int, y_tolerance: float = 4.0) -> list[TemplateCell]:
    words = page.get_text("words")
    if not words:
        from core.pdf_utils import get_page_rows

        cells: list[TemplateCell] = []
        for row_index, row in enumerate(get_page_rows(page, y_tolerance=y_tolerance)):
            for column_index, (x, text) in enumerate(row):
                cells.append(TemplateCell(
                    page_index=page_index,
                    row_index=row_index,
                    column_index=column_index,
                    x=float(x),
                    y=float(row_index),
                    text=text,
                ))
        return cells

    rows: list[list[tuple[float, float, str]]] = []
    current: list[tuple[float, float, str]] = []
    current_y = sorted(words, key=lambda w: w[1])[0][1]

    for word in sorted(words, key=lambda w: w[1]):
        x0, y0, text = word[0], word[1], word[4]
        if abs(y0 - current_y) <= y_tolerance:
            current.append((x0, y0, text))
        else:
            rows.append(sorted(current, key=lambda item: item[0]))
            current = [(x0, y0, text)]
            current_y = y0

    if current:
        rows.append(sorted(current, key=lambda item: item[0]))

    cells: list[TemplateCell] = []
    for row_index, row in enumerate(rows):
        for column_index, (x, y, text) in enumerate(row):
            cells.append(TemplateCell(
                page_index=page_index,
                row_index=row_index,
                column_index=column_index,
                x=x,
                y=y,
                text=text,
            ))
    return cells


def export_parser_template(
    pages: list,
    output_path,
    data_start_keyword: str,
    date_re: str | None = None,
    max_pages: int = 3,
) -> str | None:
    """Generate Excel template for header_mapped parser creation.

    Returns detected date format string (e.g. "yyyy/mm/dd") or None when
    date_re was provided without a matching _AUTO_DATE_PATTERNS entry.
    Raises ValueError if data_start_keyword is not found or date detection fails.
    """
    import re as _re
    from collections import defaultdict

    all_page_cells = [
        _extract_page_cells(page, pi) for pi, page in enumerate(pages[:max_pages])
    ]

    # Find header group start
    header_start_page = None
    header_start_row_idx = None
    for pi, cells in enumerate(all_page_cells):
        for c in cells:
            if data_start_keyword in c.text:
                header_start_page = pi
                header_start_row_idx = c.row_index
                break
        if header_start_page is not None:
            break

    if header_start_page is None:
        raise ValueError(f"'{data_start_keyword}'를 PDF에서 찾을 수 없습니다.")

    # Resolve date_re
    detected_format = None
    if date_re is None:
        candidate_texts = [
            c.text for c in all_page_cells[header_start_page]
            if c.row_index > header_start_row_idx
        ]
        result = _detect_date_format(candidate_texts)
        if result is None:
            raise ValueError(
                "날짜 패턴을 자동으로 감지할 수 없습니다. 날짜 형식을 직접 입력하세요 (예: yyyy/mm/dd)."
            )
        date_re, detected_format = result
    else:
        for pattern, fmt in _AUTO_DATE_PATTERNS:
            if date_re == pattern:
                detected_format = fmt
                break

    compiled_re = _re.compile(date_re)

    # Group cells by row on the header page
    page_cells = all_page_cells[header_start_page]
    cells_by_row = defaultdict(list)
    for c in page_cells:
        cells_by_row[c.row_index].append(c)

    # Identify header group rows (from keyword row up to first data row)
    header_row_indices = []
    first_data_row_idx = None
    for row_idx in sorted(cells_by_row.keys()):
        if row_idx < header_start_row_idx:
            continue
        row_texts = [c.text for c in cells_by_row[row_idx]]
        if any(compiled_re.match(t) for t in row_texts):
            first_data_row_idx = row_idx
            break
        header_row_indices.append(row_idx)

    if not header_row_indices:
        raise ValueError("제목행 그룹을 찾을 수 없습니다.")
    if first_data_row_idx is None:
        raise ValueError("첫 번째 데이터 행을 찾을 수 없습니다.")

    # Column zones from first data row (center-aligned headers give unreliable x)
    header_cells = [c for c in page_cells if c.row_index in header_row_indices]
    first_data_cells = cells_by_row[first_data_row_idx]
    zones = _compute_x_zones(first_data_cells or header_cells, x_gap=20.0)

    # Find date column x for anchor detection
    date_header_x = next(
        (c.x for c in header_cells if data_start_keyword in c.text), None
    )

    # Collect sample transaction groups
    sample_groups = []
    current_group = []

    for pi, p_cells in enumerate(all_page_cells):
        rc = defaultdict(list)
        for c in p_cells:
            rc[c.row_index].append(c)

        for row_idx in sorted(rc.keys()):
            if pi == header_start_page and row_idx < first_data_row_idx:
                continue
            row = sorted(rc[row_idx], key=lambda c: c.x)
            row_texts = [c.text for c in row]

            is_anchor = False
            if date_header_x is not None:
                closest = min(row, key=lambda c: abs(c.x - date_header_x), default=None)
                if closest and compiled_re.match(closest.text):
                    is_anchor = True
            if not is_anchor and any(compiled_re.match(t) for t in row_texts):
                is_anchor = True

            if is_anchor:
                if current_group:
                    sample_groups.append(current_group)
                if len(sample_groups) >= 5:
                    break
                current_group = [row]
            elif current_group:
                current_group.append(row)

        if len(sample_groups) >= 5:
            break

    if current_group:
        sample_groups.append(current_group)

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF"
    meta = wb.create_sheet(META_SHEET)
    fields_ws = wb.create_sheet(FIELDS_SHEET)
    config_ws = wb.create_sheet("_config")

    ws["A1"] = "제목행(회색) 셀을 노란색으로 칠해 필드를 지정하세요. 무시할 키워드는 회색으로 칠하세요."
    ws["A1"].alignment = Alignment(wrap_text=True)

    meta.append([
        "sheet", "excel_row", "excel_col",
        "page_index", "row_index", "column_index",
        "x", "y", "text", "is_header_row",
    ])

    config_ws.append(["date_format", detected_format or ""])
    config_ws.append(["data_start_keyword", data_start_keyword])
    config_ws.sheet_state = "hidden"

    HEADER_BG = "D9D9D9"
    SAMPLE_FILLS = ["FFFFFF", "EBF3FB"]
    ANCHOR_GREEN = "70AD47"

    excel_row = 3
    max_col = 1
    first_anchor_done = False

    for row_idx in header_row_indices:
        for c in sorted(cells_by_row[row_idx], key=lambda cell: cell.x):
            excel_col = _find_zone_index(c.x, zones) + 1
            max_col = max(max_col, excel_col)
            cell = ws.cell(excel_row, excel_col, c.text)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
            meta.append([
                ws.title, excel_row, excel_col,
                c.page_index, c.row_index, c.column_index,
                c.x, c.y, c.text, True,
            ])
        excel_row += 1

    green_fill = PatternFill(fill_type="solid", fgColor=ANCHOR_GREEN)
    for group_idx, group in enumerate(sample_groups):
        bg = SAMPLE_FILLS[group_idx % 2]
        fill = PatternFill(fill_type="solid", fgColor=bg)
        for row_in_group, row in enumerate(group):
            for c in row:
                excel_col = _find_zone_index(c.x, zones) + 1
                max_col = max(max_col, excel_col)
                if not first_anchor_done and row_in_group == 0 and compiled_re.match(c.text):
                    ws.cell(excel_row, excel_col, c.text).fill = green_fill
                    first_anchor_done = True
                else:
                    ws.cell(excel_row, excel_col, c.text).fill = fill
                meta.append([
                    ws.title, excel_row, excel_col,
                    c.page_index, c.row_index, c.column_index,
                    c.x, c.y, c.text, False,
                ])
            excel_row += 1
        excel_row += 1

    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    fields_ws.append(["필드 키", "필드명"])
    fields_ws["A1"].font = Font(bold=True)
    fields_ws["B1"].font = Font(bold=True)
    for key, label in STANDARD_FIELDS.items():
        fields_ws.append([key, label])
    fields_ws.column_dimensions["A"].width = 16
    fields_ws.column_dimensions["B"].width = 16

    meta.sheet_state = "hidden"
    wb.save(output_path)
    return detected_format


def read_parser_template(path) -> TemplateAnnotations:
    wb = openpyxl.load_workbook(path)
    if META_SHEET not in wb.sheetnames:
        raise ValueError("포맷 파일 metadata 시트를 찾을 수 없습니다.")

    detected_date_format = None
    if "_config" in wb.sheetnames:
        for row in wb["_config"].iter_rows(values_only=True):
            if row and row[0] == "date_format" and row[1]:
                detected_date_format = str(row[1])

    meta_ws = wb[META_SHEET]
    metadata = {}
    for row in meta_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        sheet, excel_row, excel_col, page_index, row_index, column_index, x, y, text, is_header = row
        tc = TemplateCell(
            page_index=int(page_index),
            row_index=int(row_index),
            column_index=int(column_index),
            x=float(x),
            y=float(y),
            text=str(text or ""),
        )
        metadata[(str(sheet), int(excel_row), int(excel_col))] = (tc, bool(is_header))

    header_excel_rows = [
        er for (sh, er, _ec), (_tc, is_hdr) in metadata.items()
        if is_hdr and sh != META_SHEET
    ]
    header_start_excel_row = min(header_excel_rows) if header_excel_rows else None

    field_mappings = []
    skip_keywords = []
    seen_fields = set()

    for sheet_name in wb.sheetnames:
        if sheet_name in {META_SHEET, FIELDS_SHEET, "_config"}:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                key = (sheet_name, cell.row, cell.column)
                if key not in metadata:
                    continue
                tc, is_header_row = metadata[key]
                value = str(cell.value or tc.text or "").strip()
                if not value:
                    continue
                if is_yellow(cell):
                    if not is_header_row:
                        continue
                    standard_field = value
                    if standard_field in seen_fields:
                        continue
                    seen_fields.add(standard_field)
                    row_offset = (cell.row - header_start_excel_row) if header_start_excel_row else 0
                    field_mappings.append(AnnotatedField(
                        standard_field=standard_field,
                        row_offset=row_offset,
                        x=tc.x,
                    ))
                elif is_gray(cell):
                    skip_keywords.append(value)

    return TemplateAnnotations(
        field_mappings=field_mappings,
        skip_keywords=list(dict.fromkeys(skip_keywords)),
        detected_date_format=detected_date_format,
    )
