import os
import tempfile
import openpyxl
from core.exporter import export_to_excel


def test_export_creates_broker_sheets():
    broker_raw = {
        "삼성증권": [
            {"거래일자": "2025/11/06", "거래명": "매수", "거래금액": "113,775"},
            {"거래일자": "2025/11/07", "거래명": "매도", "거래금액": "50,000"},
        ],
        "미래에셋증권": [
            {"거래일자": "2025/11/06", "거래종류": "매수", "거래금액": "200,000"},
        ],
    }
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        export_to_excel(broker_raw, path)
        wb = openpyxl.load_workbook(path)
        assert "삼성증권" in wb.sheetnames
        assert "미래에셋증권" in wb.sheetnames
        assert "통합" not in wb.sheetnames
    finally:
        os.unlink(path)


def test_export_broker_sheet_headers_match_raw_keys():
    broker_raw = {
        "테스트증권": [
            {"거래일자": "2025/01/01", "종목명": "삼성전자", "거래금액": "100,000"},
        ]
    }
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        export_to_excel(broker_raw, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["테스트증권"]
        headers = [ws.cell(1, c).value for c in range(1, 4)]
        assert "거래일자" in headers
        assert "종목명" in headers
        assert "거래금액" in headers
    finally:
        os.unlink(path)


def test_export_broker_sheet_headers_include_keys_from_all_rows():
    broker_raw = {
        "테스트증권": [
            {"거래일자": "2025/01/01", "종목명": "삼성전자"},
            {"거래일자": "2025/01/02", "거래금액": "200,000"},
        ]
    }
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        export_to_excel(broker_raw, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["테스트증권"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert headers == ["거래일자", "종목명", "거래금액"]
    finally:
        os.unlink(path)


def test_export_broker_sheet_data_rows():
    broker_raw = {
        "테스트증권": [
            {"거래일자": "2025/01/01", "거래금액": "100,000"},
            {"거래일자": "2025/01/02", "거래금액": "200,000"},
        ]
    }
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        export_to_excel(broker_raw, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["테스트증권"]
        assert ws.max_row == 3  # header + 2 data rows
    finally:
        os.unlink(path)


def test_export_uses_user_display_field_names(tmp_path):
    broker_raw = {
        "테스트증권": [
            {
                "내가쓴거래일자": "2026/05/05",
                "내가쓴종목명": "삼성전자",
                "내가쓴거래금액": "1,000",
            }
        ]
    }
    output = tmp_path / "result.xlsx"
    export_to_excel(broker_raw, str(output))

    wb = openpyxl.load_workbook(output)
    ws = wb["테스트증권"]
    headers = [ws.cell(1, col).value for col in range(1, 4)]

    assert headers == ["내가쓴거래일자", "내가쓴종목명", "내가쓴거래금액"]


def test_export_skips_empty_broker():
    broker_raw = {
        "빈증권": [],
        "테스트증권": [{"거래일자": "2025/01/01"}],
    }
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        export_to_excel(broker_raw, path)
        wb = openpyxl.load_workbook(path)
        assert "빈증권" not in wb.sheetnames
        assert "테스트증권" in wb.sheetnames
    finally:
        os.unlink(path)
