import openpyxl
from openpyxl.styles import PatternFill


def test_read_parser_template_extracts_yellow_header_fields(tmp_path):
    import openpyxl
    from openpyxl.styles import PatternFill
    from core.parser_template import META_SHEET, read_parser_template

    path = tmp_path / "format.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF"
    meta = wb.create_sheet(META_SHEET)
    meta.append(["sheet","excel_row","excel_col","page_index","row_index","column_index","x","y","text","is_header_row"])

    ws["A1"] = "거래일자"
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    meta.append(["PDF", 1, 1, 0, 3, 0, 50.0, 48.0, "거래일자", True])

    ws["B1"] = "종목명"
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    meta.append(["PDF", 1, 2, 0, 3, 1, 200.0, 48.0, "종목명", True])

    ws["C1"] = "합계"
    ws["C1"].fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
    meta.append(["PDF", 1, 3, 0, 3, 2, 300.0, 48.0, "합계", True])

    # Yellow on data row → must be IGNORED
    ws["A2"] = "2024/01/01"
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    meta.append(["PDF", 2, 1, 0, 4, 0, 50.0, 60.0, "2024/01/01", False])

    wb.save(path)
    annotations = read_parser_template(path)

    assert len(annotations.field_mappings) == 2
    date_fm = next(fm for fm in annotations.field_mappings if fm.standard_field == "거래일자")
    name_fm = next(fm for fm in annotations.field_mappings if fm.standard_field == "종목명")
    assert date_fm.x == 50.0
    assert date_fm.row_offset == 0
    assert name_fm.x == 200.0
    assert annotations.skip_keywords == ["합계"]


def test_read_parser_template_multi_header_row_offsets(tmp_path):
    import openpyxl
    from openpyxl.styles import PatternFill
    from core.parser_template import META_SHEET, read_parser_template

    path = tmp_path / "format.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF"
    meta = wb.create_sheet(META_SHEET)
    meta.append(["sheet","excel_row","excel_col","page_index","row_index","column_index","x","y","text","is_header_row"])

    # excel_row=1: 거래일자 → row_offset=0
    ws["A1"] = "거래일자"
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    meta.append(["PDF", 1, 1, 0, 3, 0, 50.0, 48.0, "거래일자", True])

    # excel_row=2: 종목명 → row_offset=1
    ws["B2"] = "종목명"
    ws["B2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    meta.append(["PDF", 2, 2, 0, 4, 1, 150.0, 60.0, "종목명", True])

    wb.save(path)
    annotations = read_parser_template(path)

    date_fm = next(fm for fm in annotations.field_mappings if fm.standard_field == "거래일자")
    name_fm = next(fm for fm in annotations.field_mappings if fm.standard_field == "종목명")
    assert date_fm.row_offset == 0
    assert name_fm.row_offset == 1


def test_read_parser_template_reads_config_sheet(tmp_path):
    import openpyxl
    from core.parser_template import META_SHEET, read_parser_template

    path = tmp_path / "format.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF"
    meta = wb.create_sheet(META_SHEET)
    meta.append(["sheet","excel_row","excel_col","page_index","row_index","column_index","x","y","text","is_header_row"])
    cfg = wb.create_sheet("_config")
    cfg.append(["date_format", "yyyy/mm/dd"])
    cfg.append(["data_start_keyword", "거래일자"])
    wb.save(path)

    annotations = read_parser_template(path)
    assert annotations.detected_date_format == "yyyy/mm/dd"


def test_read_parser_template_custom_field_name(tmp_path):
    """Unknown header text is kept as-is (custom field key)."""
    import openpyxl
    from openpyxl.styles import PatternFill
    from core.parser_template import META_SHEET, read_parser_template

    path = tmp_path / "format.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF"
    meta = wb.create_sheet(META_SHEET)
    meta.append(["sheet","excel_row","excel_col","page_index","row_index","column_index","x","y","text","is_header_row"])

    ws["A1"] = "특수컬럼명XYZ"
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    meta.append(["PDF", 1, 1, 0, 3, 0, 50.0, 48.0, "특수컬럼명XYZ", True])

    wb.save(path)
    annotations = read_parser_template(path)
    assert annotations.field_mappings[0].standard_field == "특수컬럼명XYZ"


def test_infer_standard_field_accepts_common_labels():
    from core.parser_template import infer_standard_field

    assert infer_standard_field("종목명") == "name"
    assert infer_standard_field("거래금액") == "amount"
    assert infer_standard_field("ticker") == "ticker"
    assert infer_standard_field("알수없음") is None


def test_compute_x_zones_groups_nearby_x_coordinates():
    from core.parser_template import _compute_x_zones, TemplateCell

    cells = [
        TemplateCell(0, 0, 0, x=50.0, y=10, text="거래일자"),
        TemplateCell(0, 0, 1, x=200.0, y=10, text="종목명"),
        TemplateCell(0, 0, 2, x=350.0, y=10, text="수량"),
        TemplateCell(0, 1, 0, x=52.0, y=20, text="2024/01/01"),   # same zone as 50
        TemplateCell(0, 1, 1, x=200.0, y=20, text="삼성"),         # same zone as 200
        TemplateCell(0, 1, 2, x=222.0, y=20, text="전자"),         # extra word: new zone
        TemplateCell(0, 1, 3, x=350.0, y=20, text="100"),          # same zone as 350
    ]
    zones = _compute_x_zones(cells)

    assert len(zones) == 4  # ~50, ~200, ~222, ~350


def test_find_zone_index_returns_nearest_zone_index():
    from core.parser_template import _find_zone_index

    zones = [50.0, 200.0, 350.0]
    assert _find_zone_index(49.0, zones) == 0
    assert _find_zone_index(201.0, zones) == 1
    assert _find_zone_index(340.0, zones) == 2


def test_date_format_to_re_slash():
    from core.parser_template import date_format_to_re
    assert date_format_to_re("yyyy/mm/dd") == r"\d{4}/\d{2}/\d{2}"


def test_date_format_to_re_dash():
    from core.parser_template import date_format_to_re
    assert date_format_to_re("yyyy-mm-dd") == r"\d{4}-\d{2}-\d{2}"


def test_date_format_to_re_dot():
    from core.parser_template import date_format_to_re
    assert date_format_to_re("yyyy.mm.dd") == r"\d{4}\.\d{2}\.\d{2}"


def test_date_format_to_re_two_digit_year():
    from core.parser_template import date_format_to_re
    assert date_format_to_re("yy/mm/dd") == r"\d{2}/\d{2}/\d{2}"


def test_detect_date_format_slash():
    from core.parser_template import _detect_date_format
    result = _detect_date_format(["계좌번호", "1234", "2025/11/06", "매도"])
    assert result is not None
    assert result[0] == r"\d{4}/\d{2}/\d{2}"
    assert result[1] == "yyyy/mm/dd"


def test_detect_date_format_returns_none():
    from core.parser_template import _detect_date_format
    assert _detect_date_format(["계좌번호", "ABC", "테스트"]) is None


def test_export_parser_template_creates_header_rows_with_is_header_flag(tmp_path):
    """Header group cells must have is_header_row=True in metadata."""
    import fitz
    import openpyxl
    from unittest.mock import MagicMock, patch
    from core.parser_template import META_SHEET, export_parser_template, TemplateCell

    fake_cells = [
        TemplateCell(0, 0, 0, x=50.0,  y=10.0, text="거래일자"),
        TemplateCell(0, 0, 1, x=200.0, y=10.0, text="종목명"),
        TemplateCell(0, 1, 0, x=50.0,  y=20.0, text="2024/01/01"),
        TemplateCell(0, 1, 1, x=200.0, y=20.0, text="삼성전자"),
        TemplateCell(0, 2, 0, x=50.0,  y=30.0, text="2024/01/02"),
        TemplateCell(0, 2, 1, x=200.0, y=30.0, text="카카오"),
    ]

    mock_page = MagicMock(spec=fitz.Page)
    out = tmp_path / "out.xlsx"

    with patch("core.parser_template._extract_page_cells", return_value=fake_cells):
        result = export_parser_template(
            [mock_page], out,
            data_start_keyword="거래일자",
            date_re=r"\d{4}/\d{2}/\d{2}",
        )

    wb = openpyxl.load_workbook(out)
    meta = wb[META_SHEET]
    rows = list(meta.iter_rows(min_row=2, values_only=True))

    header_rows = [r for r in rows if r[9] is True]
    data_rows   = [r for r in rows if r[9] is False]

    assert len(header_rows) == 2   # 거래일자, 종목명
    assert len(data_rows)   >= 2   # sample data rows

    header_texts = {r[8] for r in header_rows}
    assert "거래일자" in header_texts
    assert "종목명" in header_texts


def test_export_parser_template_excludes_pre_header_rows(tmp_path):
    """Rows before data_start_keyword must not appear in Excel."""
    import fitz
    import openpyxl
    from unittest.mock import MagicMock, patch
    from core.parser_template import META_SHEET, export_parser_template, TemplateCell

    fake_cells = [
        TemplateCell(0, 0, 0, x=50.0, y=5.0,  text="계좌번호: 1234-5678"),
        TemplateCell(0, 1, 0, x=50.0, y=15.0, text="거래일자"),
        TemplateCell(0, 1, 1, x=200.0,y=15.0, text="종목명"),
        TemplateCell(0, 2, 0, x=50.0, y=25.0, text="2024/01/01"),
        TemplateCell(0, 2, 1, x=200.0,y=25.0, text="삼성전자"),
    ]
    mock_page = MagicMock(spec=fitz.Page)
    out = tmp_path / "out.xlsx"

    with patch("core.parser_template._extract_page_cells", return_value=fake_cells):
        export_parser_template(
            [mock_page], out,
            data_start_keyword="거래일자",
            date_re=r"\d{4}/\d{2}/\d{2}",
        )

    wb = openpyxl.load_workbook(out)
    meta = wb[META_SHEET]
    all_texts = {r[8] for r in meta.iter_rows(min_row=2, values_only=True) if r[8]}
    assert "계좌번호: 1234-5678" not in all_texts


def test_export_parser_template_writes_config_sheet(tmp_path):
    """_config sheet must store date_format and data_start_keyword."""
    import fitz
    import openpyxl
    from unittest.mock import MagicMock, patch
    from core.parser_template import export_parser_template, TemplateCell

    fake_cells = [
        TemplateCell(0, 0, 0, x=50.0, y=10.0, text="거래일자"),
        TemplateCell(0, 0, 1, x=200.0,y=10.0, text="종목명"),
        TemplateCell(0, 1, 0, x=50.0, y=20.0, text="2024/01/01"),
        TemplateCell(0, 1, 1, x=200.0,y=20.0, text="삼성전자"),
    ]
    mock_page = MagicMock(spec=fitz.Page)
    out = tmp_path / "out.xlsx"

    with patch("core.parser_template._extract_page_cells", return_value=fake_cells):
        export_parser_template(
            [mock_page], out,
            data_start_keyword="거래일자",
            date_re=r"\d{4}/\d{2}/\d{2}",
        )

    wb = openpyxl.load_workbook(out)
    assert "_config" in wb.sheetnames
    config = {r[0]: r[1] for r in wb["_config"].iter_rows(values_only=True) if r[0]}
    assert config["data_start_keyword"] == "거래일자"


def test_export_parser_template_autodetects_date_format(tmp_path):
    """When date_re is None, auto-detection must succeed and return format string."""
    import fitz
    from unittest.mock import MagicMock, patch
    from core.parser_template import export_parser_template, TemplateCell

    fake_cells = [
        TemplateCell(0, 0, 0, x=50.0, y=10.0, text="거래일자"),
        TemplateCell(0, 1, 0, x=50.0, y=20.0, text="2024-01-01"),
        TemplateCell(0, 1, 1, x=200.0,y=20.0, text="삼성전자"),
    ]
    mock_page = MagicMock(spec=fitz.Page)
    out = tmp_path / "out.xlsx"

    with patch("core.parser_template._extract_page_cells", return_value=fake_cells):
        fmt = export_parser_template([mock_page], out, data_start_keyword="거래일자")

    assert fmt == "yyyy-mm-dd"


def test_export_parser_template_raises_when_keyword_not_found(tmp_path):
    import fitz
    from unittest.mock import MagicMock, patch
    from core.parser_template import export_parser_template, TemplateCell
    import pytest

    fake_cells = [TemplateCell(0, 0, 0, x=50.0, y=10.0, text="전혀다른내용")]
    mock_page = MagicMock(spec=fitz.Page)
    out = tmp_path / "out.xlsx"

    with patch("core.parser_template._extract_page_cells", return_value=fake_cells):
        with pytest.raises(ValueError, match="거래일자"):
            export_parser_template([mock_page], out, data_start_keyword="거래일자",
                                   date_re=r"\d{4}/\d{2}/\d{2}")
