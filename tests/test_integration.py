import tempfile, os
from collections import defaultdict
import openpyxl
from core.loader import load_pdf
from core.detector import detect_parser
from core.exporter import export_to_excel
from core.models import STANDARD_FIELDS


def _run_pipeline(pdf_path, password, tmp_path):
    pages = load_pdf(str(pdf_path), password)
    parser_class = detect_parser(pages)
    assert parser_class is not None, f"Could not detect parser for {pdf_path}"
    parser = parser_class()
    transactions, raw_rows = parser.parse(pages)
    broker_raw = defaultdict(list)
    broker_raw[parser_class.BROKER_NAME].extend(raw_rows)
    export_to_excel(
        transactions=transactions,
        broker_raw=dict(broker_raw),
        selected_fields=list(STANDARD_FIELDS.keys()),
        output_path=tmp_path,
    )
    return transactions, tmp_path


def test_samsung_full_pipeline(samsung_pdf, pdf_password):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        transactions, out = _run_pipeline(samsung_pdf, pdf_password, path)
        assert len(transactions) > 0
        wb = openpyxl.load_workbook(out)
        assert "통합" in wb.sheetnames
        assert "삼성증권" in wb.sheetnames
        ws = wb["통합"]
        assert ws.max_row > 1  # header + at least 1 data row
    finally:
        os.unlink(path)


def test_mirae_full_pipeline(mirae_pdf, pdf_password):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        transactions, out = _run_pipeline(mirae_pdf, pdf_password, path)
        assert len(transactions) > 0
        wb = openpyxl.load_workbook(out)
        assert "통합" in wb.sheetnames
        assert "미래에셋증권" in wb.sheetnames
    finally:
        os.unlink(path)


def test_combined_pipeline(samsung_pdf, mirae_pdf, pdf_password):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        all_txs = []
        broker_raw: dict = defaultdict(list)
        for pdf, pw in [(samsung_pdf, pdf_password), (mirae_pdf, pdf_password)]:
            pages = load_pdf(str(pdf), pw)
            pc = detect_parser(pages)
            txs, raws = pc().parse(pages)
            all_txs.extend(txs)
            broker_raw[pc.BROKER_NAME].extend(raws)
        export_to_excel(all_txs, dict(broker_raw), list(STANDARD_FIELDS.keys()), path)
        wb = openpyxl.load_workbook(path)
        assert "통합" in wb.sheetnames
        assert "삼성증권" in wb.sheetnames
        assert "미래에셋증권" in wb.sheetnames
        ws = wb["통합"]
        assert ws.max_row > len(all_txs)  # header row + data rows
    finally:
        os.unlink(path)
